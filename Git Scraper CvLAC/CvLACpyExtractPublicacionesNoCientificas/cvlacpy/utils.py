import openpyxl

HEADER_INPUT = ("documento","nombres","apellidos","cvlac")
def getCvlacsDirs(filename):
    """
    Obtiene la lista de los url de los cvlas a explorar
    """
    wb = openpyxl.load_workbook(filename, read_only=True )
    # ws - work sheet: hoja uno
    ws= wb[wb.sheetnames[0]]
    # Verifica el encabezado
    header = [ ws[col+"1"].value for col in "ABCD"]
    header = [ s.lower() for s in header if isinstance(s,str)]
    header = tuple(header)
    
    if header != HEADER_INPUT:
        s= " El encabezado del excel de entrada no es correcto."
        s+= "\n Debe contener: "
        for i, h in enumerate(HEADER_INPUT):
            if i> 0:
                s+=", "
            s+= h
        s+="."
        raise Exception(s)
    #Extrae la información
    datos=[list(header)]
    for r, row in enumerate(ws.iter_rows(min_row=2)):
        fila = []
        for c, cell in enumerate(row):
            fila.append(cell.value)
        datos.append(fila)
    return datos
    



